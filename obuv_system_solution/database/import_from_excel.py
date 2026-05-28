"""Импорт данных из файлов Приложения 2.

Использование:
    python database/import_from_excel.py --dir resources/import

Скрипт ожидает, что из архива Прил_2_ОЗ_КОД 09.02.07-2-2026-М1.rar
в папку resources/import будут распакованы файлы:
    Tovar.xlsx
    user_import.xlsx
    Заказ_import.xlsx
    Пункты выдачи_import.xlsx

Импортер написан устойчиво: названия колонок нормализуются, поэтому он подходит
для русских и смешанных заголовков из демонстрационных заданий.
"""

from __future__ import annotations

import argparse
import hashlib
import shutil
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.config import IMAGES_DIR, RESOURCES_DIR
from app.db import ensure_database, get_connection, get_or_create_id, password_hash

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


ALIASES = {
    'article': {'артикул', 'артикултовара', 'кодтовара', 'idтовара', 'id', 'productid'},
    'name': {'наименование', 'наименованиетовара', 'название', 'товар'},
    'category': {'категория', 'категориятовара'},
    'description': {'описание', 'описаниетовара'},
    'manufacturer': {'производитель'},
    'supplier': {'поставщик'},
    'price': {'цена', 'стоимость'},
    'unit': {'единицаизмерения', 'единица', 'единицаизм'},
    'quantity': {'количествонаскладе', 'количество', 'остаток', 'остатки'},
    'discount': {'скидка', 'действующаяскидка', 'размерскидки'},
    'image': {'изображение', 'фото', 'картинка', 'путькизображению'},
    'login': {'логин', 'login'},
    'password': {'пароль', 'password'},
    'role': {'роль', 'role'},
    'last_name': {'фамилия'},
    'first_name': {'имя'},
    'patronymic': {'отчество'},
    'order_article': {'артикулзаказа', 'номерзаказа', 'заказ', 'orderid'},
    'status': {'статус', 'статусзаказа'},
    'pickup_address': {'адрес', 'адреспунктавыдачи', 'пунктвыдачи'},
    'order_date': {'датазаказа'},
    'delivery_date': {'датавыдачи', 'датапоставки'},
}

ROLE_MAP = {
    'клиент': 'client',
    'client': 'client',
    'менеджер': 'manager',
    'manager': 'manager',
    'администратор': 'admin',
    'admin': 'admin',
}


def norm(value: Any) -> str:
    return ''.join(str(value or '').strip().lower().replace('ё', 'е').split())


def clean(value: Any) -> str:
    return str(value or '').strip()


def rows_from_xlsx(path: Path) -> list[dict]:
    if load_workbook is None:
        raise RuntimeError('Для импорта Excel установите зависимость: pip install openpyxl')
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [norm(cell) for cell in rows[0]]
    mapped_headers: list[str | None] = []
    for header in headers:
        mapped = None
        for field, variants in ALIASES.items():
            if header in variants:
                mapped = field
                break
        mapped_headers.append(mapped)

    result = []
    for row in rows[1:]:
        item = {}
        for index, value in enumerate(row):
            if index < len(mapped_headers) and mapped_headers[index]:
                item[mapped_headers[index]] = value
        if any(value is not None and str(value).strip() for value in item.values()):
            result.append(item)
    return result


def find_file(import_dir: Path, *names: str) -> Path | None:
    """Ищет Excel-файл в папке импорта рекурсивно.

    В исходном RAR все файлы лежат внутри каталога `import/`. Если распаковать
    архив прямо в `resources/import`, получится путь `resources/import/import`.
    Поэтому обычного `glob('*.xlsx')` недостаточно.
    """
    if not import_dir.exists():
        return None
    existing = {path.name.lower(): path for path in import_dir.rglob('*.xlsx')}
    for name in names:
        if name.lower() in existing:
            return existing[name.lower()]
    return None


def _case_insensitive_child(parent: Path, file_name: str) -> Path | None:
    if not parent.exists():
        return None
    target = file_name.lower()
    for child in parent.iterdir():
        if child.is_file() and child.name.lower() == target:
            return child
    return None


def _find_image_file(import_dir: Path, image_value: str | None, article: str, row_number: int) -> Path | None:
    """Находит изображение товара в типовых местах после распаковки RAR."""
    raw = clean(image_value).replace('\\', '/')
    names: list[str] = []

    if raw:
        raw_path = Path(raw)
        direct_candidates = []
        if raw_path.is_absolute():
            direct_candidates.append(raw_path)
        else:
            direct_candidates.extend([
                import_dir / raw_path,
                import_dir / 'import' / raw_path,
                RESOURCES_DIR / raw_path,
                IMAGES_DIR / raw_path,
            ])
        for candidate in direct_candidates:
            if candidate.exists() and candidate.is_file():
                return candidate
        if raw_path.name:
            names.append(raw_path.name)

    article_clean = clean(article)
    if article_clean:
        names.extend([article_clean, f'{article_clean}.jpg', f'{article_clean}.jpeg', f'{article_clean}.png'])
    names.extend([f'{row_number}.jpg', f'{row_number}.jpeg', f'{row_number}.png'])

    # Сначала быстрый поиск в наиболее частых папках.
    search_dirs = [IMAGES_DIR, import_dir, import_dir / 'import', RESOURCES_DIR]
    for name in names:
        if not name:
            continue
        possible_names = [name]
        if '.' not in Path(name).name:
            possible_names += [f'{name}.jpg', f'{name}.jpeg', f'{name}.png']
        for directory in search_dirs:
            for file_name in possible_names:
                found = _case_insensitive_child(directory, file_name)
                if found:
                    return found

    # Потом рекурсивный поиск по имени в папке импорта.
    lowered = {name.lower() for name in names if name}
    expanded = set(lowered)
    for name in list(lowered):
        if '.' not in Path(name).name:
            expanded.update({f'{name}.jpg', f'{name}.jpeg', f'{name}.png'})
    if import_dir.exists():
        for child in import_dir.rglob('*'):
            if child.is_file() and child.name.lower() in expanded:
                return child
    return None


def normalize_import_image_path(import_dir: Path, image_value: str | None, article: str, row_number: int) -> str | None:
    """Копирует фото товара в resources/images и возвращает путь для БД."""
    found = _find_image_file(import_dir, image_value, article, row_number)
    if not found:
        return clean(image_value) or None

    IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    target = IMAGES_DIR / found.name
    if found.resolve() != target.resolve():
        # Если файл с таким именем уже есть, не затираем его другим изображением.
        if target.exists() and target.stat().st_size != found.stat().st_size:
            target = IMAGES_DIR / f'{found.stem}_{abs(hash(str(found.resolve())))}{found.suffix.lower()}'
        if not target.exists():
            shutil.copy2(found, target)

    return str(target.relative_to(RESOURCES_DIR)).replace('\\', '/')


def import_products(path: Path) -> int:
    rows = rows_from_xlsx(path)
    import_dir = path.parent
    count = 0
    with get_connection() as conn:
        for row_number, row in enumerate(rows, start=1):
            article = clean(row.get('article')) or f"AUTO-{count + 1:04d}"
            name = clean(row.get('name')) or 'Без названия'
            category_id = get_or_create_id(conn, 'categories', 'category_id', 'category_name', clean(row.get('category')) or 'Не указано')
            manufacturer_id = get_or_create_id(conn, 'manufacturers', 'manufacturer_id', 'manufacturer_name', clean(row.get('manufacturer')) or 'Не указано')
            supplier_id = get_or_create_id(conn, 'suppliers', 'supplier_id', 'supplier_name', clean(row.get('supplier')) or 'Не указано')
            unit_id = get_or_create_id(conn, 'units', 'unit_id', 'unit_name', clean(row.get('unit')) or 'пара')
            price = float(row.get('price') or 0)
            quantity = int(float(row.get('quantity') or 0))
            discount = float(row.get('discount') or 0)
            image_path = normalize_import_image_path(import_dir, clean(row.get('image')), article, row_number)
            conn.execute(
                '''
                INSERT INTO products(article, product_name, category_id, description, manufacturer_id,
                                     supplier_id, price, unit_id, quantity, discount_percent, image_path)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(article) DO UPDATE SET
                    product_name = excluded.product_name,
                    category_id = excluded.category_id,
                    description = excluded.description,
                    manufacturer_id = excluded.manufacturer_id,
                    supplier_id = excluded.supplier_id,
                    price = excluded.price,
                    unit_id = excluded.unit_id,
                    quantity = excluded.quantity,
                    discount_percent = excluded.discount_percent,
                    image_path = excluded.image_path,
                    updated_at = CURRENT_TIMESTAMP
                ''',
                (
                    article, name, category_id, clean(row.get('description')), manufacturer_id,
                    supplier_id, price, unit_id, quantity, discount, image_path,
                ),
            )
            count += 1
        conn.commit()
    return count


def import_users(path: Path) -> int:
    rows = rows_from_xlsx(path)
    count = 0
    with get_connection() as conn:
        for row in rows:
            login = clean(row.get('login'))
            if not login:
                continue
            role_text = norm(row.get('role'))
            role_code = ROLE_MAP.get(role_text, 'client')
            role = conn.execute('SELECT role_id FROM roles WHERE role_code = ?', (role_code,)).fetchone()
            if not role:
                continue
            password = clean(row.get('password')) or '123456'
            conn.execute(
                '''
                INSERT INTO users(role_id, login, password_hash, last_name, first_name, patronymic)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(login) DO UPDATE SET
                    role_id = excluded.role_id,
                    password_hash = excluded.password_hash,
                    last_name = excluded.last_name,
                    first_name = excluded.first_name,
                    patronymic = excluded.patronymic
                ''',
                (
                    role['role_id'], login, password_hash(password),
                    clean(row.get('last_name')) or 'Пользователь',
                    clean(row.get('first_name')) or login,
                    clean(row.get('patronymic')),
                ),
            )
            count += 1
        conn.commit()
    return count


def import_pickup_points(path: Path) -> int:
    rows = rows_from_xlsx(path)
    count = 0
    with get_connection() as conn:
        for row in rows:
            address = clean(row.get('pickup_address')) or clean(row.get('address'))
            if not address:
                address = clean(next(iter(row.values()), ''))
            if address:
                get_or_create_id(conn, 'pickup_points', 'pickup_point_id', 'address', address)
                count += 1
        conn.commit()
    return count


def import_orders(path: Path) -> int:
    rows = rows_from_xlsx(path)
    count = 0
    with get_connection() as conn:
        for row in rows:
            article = clean(row.get('order_article')) or f"ORD-AUTO-{count + 1:04d}"
            status_id = get_or_create_id(conn, 'order_statuses', 'status_id', 'status_name', clean(row.get('status')) or 'Новый')
            pickup_id = get_or_create_id(conn, 'pickup_points', 'pickup_point_id', 'address', clean(row.get('pickup_address')) or 'Не указан')
            conn.execute(
                '''
                INSERT INTO orders(order_article, status_id, pickup_point_id, order_date, delivery_date)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(order_article) DO UPDATE SET
                    status_id = excluded.status_id,
                    pickup_point_id = excluded.pickup_point_id,
                    order_date = excluded.order_date,
                    delivery_date = excluded.delivery_date,
                    updated_at = CURRENT_TIMESTAMP
                ''',
                (
                    article, status_id, pickup_id,
                    clean(row.get('order_date')) or '2026-01-01',
                    clean(row.get('delivery_date')) or None,
                ),
            )
            count += 1
        conn.commit()
    return count


def run_import(import_dir: Path) -> None:
    ensure_database()
    product_file = find_file(import_dir, 'Tovar.xlsx')
    user_file = find_file(import_dir, 'user_import.xlsx')
    pickup_file = find_file(import_dir, 'Пункты выдачи_import.xlsx')
    order_file = find_file(import_dir, 'Заказ_import.xlsx')

    if product_file:
        print(f'Импорт товаров: {import_products(product_file)}')
    if user_file:
        print(f'Импорт пользователей: {import_users(user_file)}')
    if pickup_file:
        print(f'Импорт пунктов выдачи: {import_pickup_points(pickup_file)}')
    if order_file:
        print(f'Импорт заказов: {import_orders(order_file)}')

    if not any([product_file, user_file, pickup_file, order_file]):
        print('Файлы импорта не найдены. Проверьте папку resources/import.')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dir', type=Path, default=ROOT / 'resources' / 'import')
    args = parser.parse_args()
    run_import(args.dir)
